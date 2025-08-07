import csv
import io
from datetime import datetime
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from ..models import Product, InventoryMovement

async def export_products_to_csv(
    db: AsyncSession,
    filters: dict = None
):
    query = select(Product).where(Product.is_active == True)
    
    if filters:
        if filters.get("low_stock"):
            query = query.where(Product.stock <= Product.stock_min)
        if filters.get("category"):
            query = query.where(Product.category == filters["category"])
    
    result = await db.execute(query)
    products = result.scalars().all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Encabezados en español
    writer.writerow([
        "Código", "Nombre (ES)", "Nombre (HT)", "Precio HTG", 
        "Precio USD", "Stock", "Stock Mínimo", "Categoría"
    ])
    
    for p in products:
        writer.writerow([
            p.barcode,
            p.name.get("es", ""),
            p.name.get("ht", ""),
            str(p.price_htg),
            str(p.price_usd),
            str(p.stock),
            str(p.stock_min),
            p.category or ""
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=productos_{datetime.now().date()}.csv"
        }
    )

async def export_movements_to_excel(
    db: AsyncSession,
    filters: dict = None
):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    query = select(InventoryMovement).order_by(InventoryMovement.created_at.desc())
    
    if filters:
        if filters.get("start_date"):
            query = query.where(InventoryMovement.created_at >= filters["start_date"])
        if filters.get("end_date"):
            query = query.where(InventoryMovement.created_at <= filters["end_date"])
        if filters.get("type"):
            query = query.where(InventoryMovement.type == filters["type"])
    
    result = await db.execute(query)
    movements = result.scalars().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Encabezados
    headers = [
        "Fecha", "Producto", "Código", "Tipo", 
        "Cantidad", "Usuario", "Referencia"
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Datos
    for row_num, mov in enumerate(movements, 2):
        ws.cell(row=row_num, column=1, value=mov.created_at)
        ws.cell(row=row_num, column=2, value=mov.product.name.get("es", ""))
        ws.cell(row=row_num, column=3, value=mov.product.barcode)
        ws.cell(row=row_num, column=4, value=mov.type)
        ws.cell(row=row_num, column=5, value=mov.quantity)
        ws.cell(row=row_num, column=6, value=mov.user.email)
        ws.cell(row=row_num, column=7, value=mov.reference or "")
    
    # Ajustar ancho columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=movimientos_{datetime.now().date()}.xlsx"
        }
    )
